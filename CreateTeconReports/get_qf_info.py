import copy
import os
import sqlite3
from dataclasses import dataclass
from enum import Enum

import firebirdsql
import openpyxl


class FindPath:

    def __init__(self, path_tecon_archive):
        self.path_tecon_archive = path_tecon_archive
        self.exeption_months = []
        self.exeption_days = []

    @staticmethod
    def get_arh_path(path):
        list_path = path.split('\\')
        return f"""{path}\\archive_{list_path[-2]}{list_path[-1]}_0000_001.arc"""

    def find_path_arch(self):
        month = self.find_last(self.path_tecon_archive, self.exeption_months)
        arh_path = None
        while arh_path is None:
            arh_path = self.find_last(month, self.exeption_days)
            if arh_path is None:
                self.exeption_months.append(month)
                self.exeption_days = []
                month = self.find_last(self.path_tecon_archive, self.exeption_months)
            if month is None:
                return None
        self.exeption_days.append(arh_path)
        return self.get_arh_path(arh_path)

    def find_last(self, path, exception=None, func=os.path.getctime):
        """

        :param path:
        :param exception:
        :param func: os.path.getctime - время создания os.path.getmtime - время изменения
        :return:
        """

        dir_list = [os.path.join(path, x) for x in os.listdir(path)]
        if dir_list:
            date_list = [[x, func(x)] for x in dir_list if x not in exception]
            sort_date_list = sorted(date_list, key=lambda x: x[1], reverse=True)
            if sort_date_list:
                return sort_date_list[0][0]
            else:
                return None

    @staticmethod
    def find_last_db(path, extension='gdb'):
        dir_list = [os.path.join(path, x) for x in os.listdir(path)]
        if dir_list:
            # Создадим список из путей к файлам и дат их создания.
            date_list = [[x, os.path.getmtime(x)] for x in dir_list if x.split('.')[-1].lower() == extension.lower()]

            # Отсортируем список по дате создания в обратном порядке
            sort_date_list = sorted(date_list, key=lambda x: x[1], reverse=True)

            if sort_date_list:
                # Выведем первый элемент списка. Он и будет самым последним по дате
                return sort_date_list[0][0]
            else:
                return None


class State(Enum):
    UNKNOWN = 0
    ARRIVE = 1
    GONE = 2


@dataclass
class StateColor:
    GONE: str = '008345'
    ARRIVE: str = 'ff8345'
    UNKNOWN: str = 'adacaa'


@dataclass
class Page:
    ID: int = 0
    PID: int = 0
    NAME: str = ''


class PagesStruct:

    def __init__(self, cursor):
        self.pages_dict_SQL = {}
        self.cursor = cursor
        self.get_all_pages()
        self.get_struct_pages()

    def get_all_pages(self):
        self.cursor.execute(f"""select ID, PID, NAME from GRPAGES""")
        self.pages_dict_SQL = {ID: Page(ID=ID, PID=PID, NAME=NAME) for ID, PID, NAME in self.cursor.fetchall()}

    def get_struct_pages(self):
        for page in self.pages_dict_SQL.values():
            while True:
                try:
                    parent = self.pages_dict_SQL[page.PID]
                    page.PID = parent.PID
                    page.NAME = f"{parent.NAME}//{page.NAME}"
                except:
                    break


class KlassStruct(PagesStruct):
    def get_all_pages(self):
        self.cursor.execute(f"""select ID, PID, NAME from KLASSIFIKATOR""")
        self.pages_dict_SQL = {ID: Page(ID=ID, PID=PID, NAME=NAME) for ID, PID, NAME in self.cursor.fetchall()}


class ObjectTypeInfo:
    """
    Класс для типа
    """

    def __init__(self, fbd_cur, object_type_name: str, object_type_chanels: list):
        """

        :param object_type_name:
        :param object_type_chanels:
        """
        self.object_type_name = object_type_name
        self.object_type_chanels = object_type_chanels
        self.object_type_id = self.get_obgect_type_id(fbd_cur)
        self.object_type_chanels_id = self.get_object_type_chanels_id(fbd_cur)

    def get_obgect_type_id(self, fbd_cur):
        select = f"""select ID from OBJTYPE  where name = '{self.object_type_name}'"""
        fbd_cur.execute(select)
        return fbd_cur.fetchall()[0][0]

    def get_object_type_chanels_id(self, fbd_cur):
        select = f"""select OBJTYPEPARAM.ID, OBJTYPEPARAM.DISC 
from OBJTYPE JOIN OBJTYPEPARAM 
ON OBJTYPE.ID = OBJTYPEPARAM.PID 
where OBJTYPE.NAME = '{self.object_type_name}' and OBJTYPEPARAM.NAME in ('{"', '".join(self.object_type_chanels)}')"""

        fbd_cur.execute(select)

        return {str(ID): DISC for ID, DISC in fbd_cur.fetchall()}


class QFSearch:
    """

    """

    def __init__(self, fdb_cur, object_types: list, id_klass: int):
        """

        :param object_types:
        :param object_chanels:
        :param id_klass:
        """
        self.list_of_object = self.get_all_marks(object_types, id_klass, fdb_cur)

    def get_all_marks(self, object_types, id_klass, fdb_cur):
        """

        :param object_types:
        :param id_klass:
        :return:
        """
        result = {}
        for object_type in object_types:
            select = f"select ID, MARKA, NAME from CARDS  where klid = {id_klass} and objtypeid = {object_type.object_type_id} "
            fdb_cur.execute(select)
            try:
                db_list = fdb_cur.fetchall()
                for id, marka, name in db_list:
                    result.update(
                        {marka: ObjectInfo(marka=marka, name=name, id=id, type_info=object_type, fdb_cur=fdb_cur)})

            except Exception as e:
                print(e)
                result = {}

        return result


class ObjectInfo:

    def __init__(self, marka, name, id, type_info, fdb_cur):
        self.marka = marka
        self.name = name
        self.id = id
        self.chanels, self.chanels_state = self.get_chanels(type_info, fdb_cur)

    def get_chanels(self, type_info, fdb_cur):
        select = f"""select objtypeparamid, id from cardparams  where cardid = {self.id} and  objtypeparamid in ('{"', '".join(type_info.object_type_chanels_id.keys())}')"""
        fdb_cur.execute(select)
        data_db = fdb_cur.fetchall()
        return {id: type_info.object_type_chanels_id[str(typeid)] for typeid, id in data_db}, {
            type_info.object_type_chanels_id[str(typeid)]: State.UNKNOWN for typeid, id in data_db}


class DBResult:

    def __init__(self, data_base_path, arh_path):

        self.find_path = FindPath(path_tecon_archive=arh_path)
        self.path_fbd = self.find_path.find_last_db(data_base_path)
        self.server = '127.0.0.1'
        self.color_state = {State.UNKNOWN: StateColor.UNKNOWN, State.ARRIVE: StateColor.ARRIVE,
                            State.GONE: StateColor.GONE}

    def get_object_type_info(self, object_types: list, chanel_list: list) -> list:
        """

        :param chanel_list:
        :param object_types:
        :return:
        """
        result_list = []
        for type, chanels in zip(object_types, chanel_list):
            result_list.append(ObjectTypeInfo(fbd_cur=self.fdb_cur, object_type_name=type, object_type_chanels=chanels))

        return result_list

    def firebird_db_init(self):
        self.fdb_conn = firebirdsql.connect(
            host=self.server,
            database=self.path_fbd,
            port=3050,
            user='sysdba',
            password='masterkey',
            charset='utf8'
        )
        self.fdb_cur = self.fdb_conn.cursor()

        # определяем классификаторы
        klid = KlassStruct(self.fdb_cur)

        object_types_id = self.get_object_type_info(object_types=['QF_0x3', 'QF_0x2', 'BOX_VV'], chanel_list=[
            ['.MsgStOn', '.MsgStOff', '.MsgStUnc', '.MsgStDbl', '.MsgStInvalid', '.MsgEOff', '.MsgInvalidEOff'],
            ['.MsgOn', '.MsgOff', '.MsgUnc', '.MsgDbl', '.MsgInvalid'],
            ['.Msg_Q_Dbl', '.Msg_Q_DU_Dbl', '.Msg_Q_DU_Invalid', '.Msg_Q_DU_Off', '.Msg_Q_DU_On', '.Msg_Q_DU_Unc',
             '.Msg_Q_Invalid', '.Msg_Q_Off', '.Msg_Q_On', '.Msg_Q_Unc', '.Msg_QS_Dbl', '.Msg_QS_Invalid', '.Msg_QS_Off',
             '.Msg_QS_On', '.Msg_QS_Test', '.Msg_QS_Unc']])

        electricity = []
        data_firebird = {}

        for klass in klid.pages_dict_SQL.keys():
            klass_data = klid.pages_dict_SQL[klass]
            if "Электроснабжение" in klass_data.NAME:
                electricity.append(klass_data)
                list_of_object = QFSearch(fdb_cur=self.fdb_cur, id_klass=klass_data.ID,
                                          object_types=object_types_id).list_of_object
                if list_of_object:
                    data_firebird.update({klass_data.NAME: list_of_object})

        self.result_data = copy.deepcopy(data_firebird)
        self.search_to_arch(data_firebird)
        self.data_to_excel()

    def column_with(self, text, collumn, list_of_wight):
        k = 1.3
        while len(list_of_wight) <= collumn - 1:
            list_of_wight.append(0)

        if list_of_wight[collumn - 1] < int(len(text) * k):
            list_of_wight[collumn - 1] = int(len(text) * k)

        return list_of_wight

    def data_to_excel(self):
        wb = openpyxl.Workbook()  # Создали книгу
        for ind_sheet, klass in enumerate(self.result_data.keys()):
            work_sheet = wb.create_sheet(klass.split('/')[-1],
                                         ind_sheet)  # Создали лист с названием и сделали его активным
            list_of_wight = []
            for id_object, object in enumerate(self.result_data[klass].keys()):
                work_sheet.cell(row=id_object + 1, column=1).value = object
                list_of_wight = self.column_with(text=object, collumn=1, list_of_wight=list_of_wight)
                collumn = 2
                for chanel in self.result_data[klass][object].chanels_state.keys():
                    if self.result_data[klass][object].chanels_state[chanel] == State.ARRIVE:
                        work_sheet.cell(row=id_object + 1, column=collumn).value = chanel
                        list_of_wight = self.column_with(text=chanel, collumn=collumn, list_of_wight=list_of_wight)
                        collumn += 1
                if collumn == 2:
                    work_sheet.cell(row=id_object + 1, column=collumn).value = 'Нет данных в архиве'
                    list_of_wight = self.column_with(text='Нет данных в архиве', collumn=collumn, list_of_wight=list_of_wight)

            for ind_coll, collumn_wight in enumerate(list_of_wight):
                work_sheet.column_dimensions[chr(65 + ind_coll)].width = collumn_wight

                # work_sheet.cell(row=id_object + 1, column=collumn).fill = PatternFill(fill_type='solid',
                #                                                                             start_color=
                #                                                                             self.color_state[
                #                                                                                 self.result_data[
                #                                                                                     klass][
                #                                                                                     object].chanels_state[
                #                                                                                     chanel]])

        wb.save('result.xlsx')

    def get_type(self, objtypeid):
        self.fdb_cur.execute(
            f"select NAME from OBJTYPE where id = {objtypeid}")
        return self.fdb_cur.fetchall()[0][0]

    def klass_enumerate(self, data_kalss, cursor):
        klass_dict = copy.deepcopy(data_kalss)
        for klass in data_kalss.keys():
            new_klass = self.objects_enumerate(data_objects=data_kalss[klass], cursor=cursor, klass=klass)
            if klass:
                klass_dict.update({klass: new_klass})
            else:
                klass_dict.pop(klass)
        return klass_dict

    def objects_enumerate(self, data_objects, cursor, klass):
        objects_dict = copy.deepcopy(data_objects)
        for object in data_objects.keys():
            new_object = self.chanels_enumerate(data_chanels=data_objects[object], cursor=cursor, object=object,
                                                klass=klass)
            if new_object:
                objects_dict.update({object: new_object})
            else:
                objects_dict.pop(object)
        return objects_dict

    def chanels_enumerate(self, data_chanels, cursor, object, klass):
        chanels_dict = copy.deepcopy(data_chanels.chanels)
        for chanel in data_chanels.chanels.keys():
            answer = self.answer_to_archive(chanel, cursor)
            if answer is not State.UNKNOWN:
                self.result_data[klass][object].chanels_state[data_chanels.chanels[chanel]] = answer
                chanels_dict.pop(chanel)
        data_chanels.chanels = chanels_dict
        if chanels_dict:
            return data_chanels
        else:
            return None

    def search_to_arch(self, data_firebird):
        path_archive = self.find_path.find_path_arch()
        conn = sqlite3.connect(path_archive)
        cursor = conn.cursor()

        while path_archive and data_firebird.keys():
            data_firebird = self.klass_enumerate(data_kalss=data_firebird, cursor=cursor)
            path_archive = self.find_path.find_path_arch()
            if path_archive:
                conn = sqlite3.connect(path_archive)
                cursor = conn.cursor()
            else:
                conn.close()

    def answer_to_archive(self, id, cursor):
        response = {'80?`\'': State.ARRIVE, '00`\'': State.GONE}
        select = f"""SELECT Data FROM ArchiveData where Tagid = {id} ORDER BY StoredTime DESC LIMIT 1"""
        cursor.execute(select)
        res = cursor.fetchall()

        if res:
            return response.get(str(res[0][0]).split('\\x')[-1], State.UNKNOWN)
        else:
            return State.UNKNOWN


if __name__ == '__main__':
    new_BD = DBResult(data_base_path='C:\\_Scada\\DB\\01 KS4 Nimnyrskay\\KS4\\06.02.21',
                      arh_path='C:\\_Scada\\Scada.Archive\\Arc\\Archive')
    new_BD.firebird_db_init()

    # test_path = FindPath(path_tecon_archive='C:\\_Scada\\Scada.Archive\\Arc\\Archive')
    # test = True
    # while test:
    #     test = test_path.find_path_arch()
    #     print(test)
