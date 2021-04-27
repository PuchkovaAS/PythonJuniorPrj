import openpyxl #Подключаем библиотеку
from openpyxl.styles import Font, Fill, PatternFill

wb = openpyxl.Workbook() #Создали книгу
work_sheet = wb.create_sheet('Test sheet', 0) #Создали лист с названием и сделали его активным
work_sheet['A1'] = 'RED'
work_sheet['A2'] = 'GREEN'
work_sheet['A1'].fill = PatternFill(fill_type='solid', start_color='ff8345')
work_sheet['A2'].fill = PatternFill(fill_type='solid', start_color='008345')
wb.save('ok.xlsx')